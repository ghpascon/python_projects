import os
import asyncio
import bleak
from openpyxl import Workbook, load_workbook

async def scan_and_save_devices():
    print("Scanning for nearby Bluetooth devices...")

    # Scan for nearby Bluetooth devices
    async with bleak.BleakScanner() as scanner:
        # Introducing a delay before scanning to give more time for device names to be fetched
        await asyncio.sleep(5)
        devices = await scanner.discover()

        # Create a new workbook if 'devices.xlsx' does not exist
        file_path = "devices.xlsx"
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Devices"
            ws.append(["Name", "Address"])
        else:
            wb = load_workbook(file_path)
            ws = wb.active

            # Clear existing data
            ws.delete_rows(2, ws.max_row)

        # Write device information to the worksheet
        for device in devices:
            ws.append([device.name or "Unknown", device.address])

        # Save the workbook
        wb.save(file_path)
        print(f"\n\nDevice information saved to '{file_path}'\n\n")

if __name__ == "__main__":
    asyncio.run(scan_and_save_devices())
