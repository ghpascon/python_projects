import pyautogui
import time
from openpyxl import load_workbook

# Carregar o arquivo Excel
caminho_arquivo = 'products_to_register.xlsx'
wb = load_workbook(caminho_arquivo)

# Selecionar a planilha desejada
planilha = wb.active

# Lista para armazenar os dados
dados = []

time.sleep(3)
pyautogui.press('win')
time.sleep(1)
pyautogui.write('edge')
pyautogui.press('enter')
time.sleep(3) 
pyautogui.moveTo(1010, 43, duration=0.3)  
pyautogui.click()  
time.sleep(0.5) 
pyautogui.click()  
pyautogui.write('http://127.0.0.1:8000/')
pyautogui.press('enter')
time.sleep(3) 

pyautogui.moveTo(185, 275, duration=0.3)  



for linha in range(2, planilha.max_row + 1):
    valores_linha = []
    pyautogui.click()  
    for coluna in range(2, planilha.max_column + 1):
        valor_celula = planilha.cell(row=linha, column=coluna).value
        valores_linha.append(valor_celula)
        pyautogui.write(str(valor_celula))  # Converta para string para garantir a escrita correta
        pyautogui.press('tab')
    dados.append(valores_linha)
    pyautogui.press('enter')
    time.sleep(1)



