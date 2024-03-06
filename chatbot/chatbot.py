#libs
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

def send_email(recipient, subject, message, attachment_name, attachment_path, sender, password, smtp_server, smtp_port):
    # Set up the email
    recipient = email
    subject = 'Gabriel Pascon Resume.'
    message = f'Hello {name}, as requested, here is my resume.'
    attachment_name = 'resume_ghp.pdf'
    attachment_path = 'resume_ghp.pdf'
    sender = my_email
    password = my_pass 
    smtp_server = 'smtp.gmail.com'
    smtp_port = 465

    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = recipient
    msg['Subject'] = subject

    # Message body
    body = message
    msg.attach(MIMEText(body, 'plain'))

    # Attach the file
    with open(attachment_path, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', "attachment; filename= %s" % attachment_name)
        msg.attach(part)

    # Connect to SMTP server and send email
    try:
        server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        server.login(sender, password)
        server.sendmail(sender, recipient, msg.as_string())
        server.quit()
        print("Email sent successfully!")
    except Exception as e:
        print("Error sending email:", str(e))

#vars
loop=0
my_name=''
my_email=''
my_phone=''
my_pass=''
name=''
email=''
recipient = 'email'
subject = ''
message = ''
attachment_name = ''
attachment_path = ''
sender = ''
password = ''  
smtp_server = ''
smtp_port = ''




#open config file
with open('config.txt', 'r') as file:
    lines = file.readlines() 

#read each line and get the parameters
for line in lines:
    param = line.strip().split(':')
    if len(param) == 2:
        parameter, value = param
        if parameter.strip() == 'name':
            my_name = value.strip()
        elif parameter.strip() == 'email':
            my_email = value.strip()
        elif parameter.strip() == 'phone':
            my_phone = value.strip()
        elif parameter.strip() == 'pass':
            my_pass = value.strip()

#open the data store
workbook = openpyxl.load_workbook('data.xlsx')

# Select the active sheet
sheet = workbook['register']

# Find the last row with content in column A
last_row = sheet.max_row

#welcome to the program and get name and email
name = input('Welcome to chatbot!!!\n\nwhat is your name\n')
email = input(f'\nHello {name}, what is your email?\n')

#write parameter on xlxs
sheet.cell(row=last_row + 1, column=1).value = name
sheet.cell(row=last_row + 1, column=2).value = email

# Save the changes to the file
workbook.save('data.xlsx')

#loop
while loop==0:
    action = input(f'\n{name}, select an option below:\n[1] - My infos.\n[2] - My skills.\n[3] - Recive my resume in your email.\n[4] - exit.\n')
    if action == '1':
        print(f'\nName: {my_name}\nPhone: {my_phone}\nEmail: {my_email}')
    elif action == '2':
        print('\n-Python\n-C++\n-embeeded software\n-esp32\n-CLP\n-Automation')
    elif action == '3':
        print(f'\nSending my resume to {email}.')
        send_email(recipient, subject, message, attachment_name, attachment_path, sender, password, smtp_server, smtp_port)
    elif action == '4':
        loop=1
        print(f'\nBye bye {name}')
    else:
        print('\nInvalid number')
 